"""Load the department rotation from the Ausbildungseinsatzplan Excel files.

The Excel files (e.g. ``Ausbildungseinsatzplan25_26.xlsx``) live in a folder
next to the reports (default: ``<AN_OUTPUT_DIR>/Ausbildungseinsatzplan``).
Each workbook has one sheet per Lehrjahr; row 1 holds the trainee names,
column A the week date, and each cell the department for that trainee and
week. Empty cells are Berufsschule blocks.

All workbooks in the folder are merged into one ``{week_monday: department}``
mapping for the configured trainee. Results are cached per file mtime, so
report generation does not re-parse unchanged files.
"""

from __future__ import annotations

from datetime import date, datetime, timedelta
from pathlib import Path

import openpyxl

from . import credentials

# Empty cells in the plan mean Berufsschule block weeks.
EMPTY_MEANS = "Berufsschule"

_cache: dict[str, tuple[tuple, dict[date, str]]] = {}


def plan_dir() -> Path:
    """Plan folder: profile/env setting, else <output_dir>/Ausbildungseinsatzplan."""
    profile = credentials.get_profile()
    if profile["einsatzplan_dir"]:
        return Path(profile["einsatzplan_dir"])
    if profile["output_dir"]:
        return Path(profile["output_dir"]) / "Ausbildungseinsatzplan"
    return Path()


def _week_monday(d: date) -> date:
    return d - timedelta(days=d.weekday())


def _find_name_column(header_row, name: str) -> int | None:
    """Locate the trainee's column: exact match first, then substring."""
    wanted = name.strip().casefold()
    cells = [(i, str(c.value).strip()) for i, c in enumerate(header_row) if c.value]
    for i, text in cells:
        if text.casefold() == wanted:
            return i
    for i, text in cells:
        if wanted in text.casefold() or text.casefold() in wanted:
            return i
    return None


def _parse_workbook(path: Path, name: str) -> dict[date, str]:
    weeks: dict[date, str] = {}
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        for sheet_name in wb.sheetnames:
            if not sheet_name.lower().startswith("lehrjahr"):
                continue
            ws = wb[sheet_name]
            rows = ws.iter_rows()
            try:
                header = next(rows)
            except StopIteration:
                continue
            col = _find_name_column(header, name)
            if col is None:
                continue
            for row in rows:
                raw_date = row[0].value if row else None
                if isinstance(raw_date, datetime):
                    day = raw_date.date()
                elif isinstance(raw_date, date):
                    day = raw_date
                else:
                    continue
                value = row[col].value if len(row) > col else None
                dept = str(value).strip() if value is not None else ""
                weeks[_week_monday(day)] = dept or EMPTY_MEANS
    finally:
        wb.close()
    return weeks


def source_files() -> list[str]:
    """The plan workbooks that would be parsed (for status/debug output)."""
    directory = plan_dir()
    if not directory.is_dir():
        return []
    return [
        p.name for p in sorted(directory.glob("*.xlsx"))
        if not p.name.startswith("~$")
    ]


def load_rotation(name: str) -> dict[date, str]:
    """Merged week->department mapping for ``name`` from all plan files.

    Returns an empty dict when the folder or files are missing/unreadable -
    callers fall back to the bundled rotation.json.
    """
    directory = plan_dir()
    if not directory.is_dir():
        return {}

    merged: dict[date, str] = {}
    for path in sorted(directory.glob("*.xlsx")):
        if path.name.startswith("~$"):  # Excel lock files
            continue
        key = str(path)
        try:
            stat = (path.stat().st_mtime_ns, path.stat().st_size)
        except OSError:
            continue
        cached = _cache.get(key)
        if cached and cached[0] == stat:
            weeks = cached[1]
        else:
            try:
                weeks = _parse_workbook(path, name)
            except Exception:
                continue
            _cache[key] = (stat, weeks)
        merged.update(weeks)
    return merged
